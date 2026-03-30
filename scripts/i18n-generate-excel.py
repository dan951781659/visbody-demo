#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成翻译压缩稿 Excel
由本地化专家人工审核并压缩了 140 条超限条目，此脚本将结果写入 Excel。

用法：
  pip install openpyxl
  python scripts/i18n-generate-excel.py [输入CSV] [输出xlsx]
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("请先安装依赖：pip install openpyxl")


# =========================================================
# 场景分类（与 i18n-compress-translation.py 保持一致）
# =========================================================
def classify(key: str) -> tuple[str, int]:
    k = key.lower()
    segs = k.split(".")
    last = segs[-1] if segs else k
    if last in ("btn", "button") or any(s in ("btn", "button") for s in segs):
        return "按钮", 15
    if last in ("title", "label", "header", "name"):
        return "标题/标签", 25
    if "placeholder" in k:
        return "占位提示", 45
    if "error" in k or last == "err":
        return "错误消息", 65
    stripped = last.rstrip("0123456789")
    if stripped in ("tip", "tips") or last == "tip":
        return "提示", 80
    if last in ("desc", "description"):
        return "描述", 120
    return "说明文本", 200


# =========================================================
# 人工审核后的压缩译文（140 条）
# 原则：以 C 列新译文为基准，保留语法正确性和准确性，控制在场景字符上限内
# =========================================================
COMPRESSED = {
    # ---------- account ----------
    "account.create.name.tips3":
        "5–100 karakter; számok, betűk és speciális jelek (@.+_-/*!#$%éüñ) megengedettek",
    "account.edit.setnewpwderror1":
        "Az új jelszó nem egyezhet meg a régivel. Adjon meg egy másikat.",

    # ---------- active ----------
    "active.model.title": "Díjköteles funkciók",

    # ---------- ai ----------
    "ai.mes":
        "Figyelem: Ez az edzési terv csak útmutató. Módosítsa egyéni körülményei alapján."
        " Kerülje a túledzést és a diéta szigorú korlátozását az egészségkárosodás elkerüléséhez.",
    "ai.msg2":
        "Figyelem: Ez a fitneszprogram csak útmutató, módosítsa személyes igényei szerint."
        " Kerülje a túlzott edzést és étrend-korlátozást a szervezet védelme érdekében.",
    "ai.title": "Kuafu AI asszisztens",
    "aim8sg":
        "Tisztelt edző, AI asszisztense vagyok. Bal oldalt láthatja a tanuló jelentéséből"
        " készített javaslatokat és edzéstervet. Módosításhoz kattintson a kommunikációs"
        " gombra – örömmel segítek!",

    # ---------- aireport ----------
    "aireport.abnormal.btn.generateoutline": "Terv készítése",
    "aireport.abnormal.conclusion.btn.showreport": "Jelentés",
    "aireport.abnormal.error5":
        "AI-elemzéshez testalak/testkerület mérés is szükséges.",
    "aireport.abnormal.error6":
        "AI-elemzés nem elérhető. Végezzen el egy testmérési csoportot is.",
    "aireport.abnormal.improvemap.title": "Fejlesztési térkép",
    "aireport.abnormal.items.title": "Testi mutatók elemzése",
    "aireport.abnormal.nodata.tip1":
        "Testalkata és izomzata rendben. Folytassa az egészséges étkezést és mozgást.",
    "aireport.abnormal.nodata.tip2":
        "Mutatói rendben. Tartsa meg egészséges életmódját és figyelje állapotát.",
    "aireport.abnormal.sport.riskchain.header": "Sportkockázat-elemzés",
    "aireport.dialog.bottom.tip1":
        "A {Kuafu*Deepseek} generálta – csak tájékoztatásul.",
    "aireport.dialog.popup.btn": "AI-tanácsadás",
    "aireport.dialog.stopchat.btn": "Leállítás",

    # ---------- api ----------
    "api.error1":
        "Ez a felhasználónév ellenőrzés alatt áll, kérjük várjon.",
    "api.error2":
        "A fióknév-ellenőrzés sikertelen. Módosítsa és küldje újra.",

    # ---------- bind ----------
    "bind.user.email.placeholder": "Adja meg a felhasználó e-mail-jét",
    "bind.user.msg.1":
        "Új felhasználóhoz tartozó e-mail-cím. Kérjük, ellenőrizze a helyességét."
        " Ha hibás, törölje és írja be újra. Ha helyes, töltse ki a felhasználói adatokat.",
    "bind.user.msg.7":
        "Új felhasználóhoz tartozó telefonszám. Kérjük, ellenőrizze a helyességét."
        " Ha hibás, törölje és írja be újra. Ha helyes, töltse ki a felhasználói adatokat.",
    "bind.user.phone.placeholder": "Adja meg a felhasználó telefonszámát",

    # ---------- bmMass ----------
    "bmMass.ECW.desc":
        "Extracelluláris folyadék: a sejten kívüli testnedvek összessége (vérplazma, szövetnedv).",
    "bmMass.VFA.desc":
        "A belső zsír a szervek között helyezkedik el és hasi elhízást okoz."
        " A zsírterület a belső zsír keresztmetszetét mutatja.",
    "bmMass.waist.mld.advice1":
        "Végezzen 1-2 alakformáló kezelést a kollagéntermelés és bőrrugalmasság javítására."
        " Fogyasszon több fehérjét és egészséges zsírt, kombinálva erőedzéssel és kardiózással.",
    "bmMass.waist.mld.advice3":
        "Hosszabb alakformáló kezelési sorozat ajánlott a test feszesítéséhez és a kollagén"
        " regenerálásához. Csökkentse a kalóriadús ételeket, növelje a rostbevitelt,"
        " és végezzen aerob és core edzéseket.",
    "bmMass.waist.mld.tip1":
        "A derék-csípő arány alacsony: alultápláltság vagy betegség kockázatát jelzi.",
    "bmMass.waist.mld.tip2":
        "A derék-csípő arány normál tartományban van, az egészségügyi kockázat alacsony.",
    "bmMass.waist.mld.tip3":
        "A derék-csípő arány a normál felett van, magasabb egészségügyi kockázattal.",

    # ---------- bsShape ----------
    "bsShape.bodySlope.desc":
        "A váll- és medenceközéppont összekötő vonala és a frontális középső vonal"
        " által bezárt szög.",

    # ---------- client ----------
    "client.form.item.contactName.placeholder":
        "A kapcsolattartó neve (vállalat/brand/üzlet)",
    "client.list.agentcode.placeholder": "Az értékesítési menedzser kódja",

    # ---------- device ----------
    "device.active.form.tip1":
        "Új eszköz észlelve – a használathoz aktiválás szükséges!",
    "device.active.modal.vip1tip.title": "Ön már Pro felhasználó",
    "device.image.tips":
        "Töltsön fel éles, teljes előnézeti képet a készülékről (max. 30 MB).",
    "device.list.importData.button.import": "Importálás most",
    "device.list.importData.button.search": "Lekérdezés",
    "device.list.importData.button.view": "Megtekintés",
    "device.list.importData.description":
        "A fiókjában <p>{count}</p> eszköz van. Válassza ki, melyikre importálja"
        " a mérési adatokat.",
    "device.list.importData.tip.error1":
        "Kérjük, válassza ki az adatimportálási eszközt.",
    "device.list.importData.title": "Mérési adatok importálása",
    "device.modal.vip1.tip":
        "Az alapcsomagra épülő verzió AI-értelmezéssel és többkörös AI-párbeszéddel.",
    "device.pay.btn.confirmEceipt": "Nyugta fogadása",
    "device.pay.btn.invoice": "Számla",
    "device.pro.tip.title": "Pro frissítés szükséges",
    "device.renew.form.tip1":
        "A Basic előfizetés hamarosan lejár. Kérjük, újítsa meg időben.",
    "device.renew.form.title": "Előfizetés megújítása",
    "device.tip.setvip1.btn": "Pro frissítés",

    # ---------- feeOrder ----------
    "feeOrder.search.placeholder": "Adja meg a rendelési számot",
    "feeOrder.search.placeholder.cid": "Adja meg az ügyfél azonosítóját",
    "feeOrder.search.placeholder1": "Rendelési szám vagy ügyfél neve",

    # ---------- forget ----------
    "forget.pwd.title": "Adja meg: {addr}",

    # ---------- hipJoint ----------
    "hipJoint.report.abuctionClus.title": "Csípőabdukció elemzése",
    "hipJoint.report.antexionClus.title": "Csípőhajlítás elemzése",

    # ---------- login ----------
    "login.auth.loginexpired.error":
        "A(z) {name} Pro lejárt. Kérjük, újítsa meg a Pro-előfizetést!",

    # ---------- message ----------
    "message.menu.training.plan.template.tips":
        "Végezze el a táblázatban lévő edzéseket, és jelölje be az elvégzett sorokat.",
    "message.menu.training.plan.template.tips2":
        "Figyelem: Az edzési terv csak ajánlás. Kerülje a túledzést és szigorú diétát.",
    "message.menu.training.question.desc":
        "Üdvözöljük! Töltse ki a kérdőívet, hogy segíthessünk állapota elemzésében."
        " Adatait bizalmasan kezeljük.",
    "message.menu.training.question.postpartum.day.placeholder17":
        "Adja meg terhesség előtti súlyát",
    "message.menu.training.question.postpartum.day.placeholder18":
        "Napi munkával/tanulással töltött idő:",
    "message.menu.training.question.postpartum.day.placeholder20":
        "B) Könnyű mozgás/állás (pl. tanár, irodista)",
    "message.menu.training.question.postpartum.day.placeholder21":
        "C) Közepes aktivitás (diák, sofőr, eladó)",
    "message.menu.training.question.postpartum.day.placeholder23":
        "Napi séta/bevásárlás/házimunka ideje:",
    "message.menu.training.question.postpartum.day.placeholder28":
        "Napi séta/bevásárlás/házimunka ideje:",
    "message.menu.training.question.postpartum.day.placeholder37":
        "Heti testmozgás ideje (munkán kívül):",
    "message.menu.training.question.postpartum.day.placeholder43":
        "A) Séta, jóga, alacsony intenzitású mozgás",
    "message.menu.training.question.postpartum.day.placeholder44":
        "B) Közepes mozgás: pilates, futás, úszás",
    "message.menu.training.question.postpartum.day.placeholder46":
        "Hogyan ingázik (több válasz lehetséges):",
    "message.menu.training.question.postpartum.day.placeholder51":
        "Van-e ilyen problémája? (több is választható)",
    "message.menu.training.question.postpartum.day.placeholder9":
        "Adja meg terhesség előtti súlyát",
    "message.menu.training.question.pregnancy.week.placeholder":
        "Adja meg terhességi hetét",
    "message.menu.training.question.title": "Kérdőív kitöltése",
    "message.menu.training.selectReport.placeholder":
        "Válasszon jelentést az edzésterv-készítéshez",
    "message.vrexp.abnormal.error6":
        "Más üzlet jelentése – ellenőrizze és próbálja újra!",

    # ---------- pay ----------
    "pay.activc.err.pay.state":
        "Fizetés folyamatban, művelet nem lehetséges! Kérjük, csak egy személy"
        " használja a rendszert a dupla fizetés elkerülése érdekében.",
    "pay.activc.pro.pricee.basic.title": "1. Számlázási képlet",
    "pay.activc.pro.pricee.pro.desc":
        "Ha a Basic lejár az upgrade előtt, csak az időarányos különbözetet kell kifizetni.",
    "pay.activc.pro.pricee.pro.title": "2. Frissítési árkülönbség",
    "pay.activc.pro.pricee.select.coupon.explain.text3":
        "3. Ha a kupon értéke meghaladja a fizetendő összeget: a különbözet nem térül vissza"
        " és nem halmozódik. A fizetendő összeg automatikusan 0-ra csökken.",
    "pay.activc.pro.pricee.select.coupon.title": "Eszközválasztás kuponhoz",
    "pay.ok.invoice.success.tip":
        "A számla 7–15 munkanapon belül érkezik az e-mail-jére.",
    "pay.tip":
        "Megjegyzés: Nem kell fizetnie, csak hagyja jóvá a szolgáltatás indítását",

    # ---------- payvip1 ----------
    "payvip1.tip":
        "Az Alapra épülő Pro: AI-értelmezés és többkörös AI-párbeszéd.",
    "payvip1.tishi":
        "A legkésőbb lejáró eszköz: [{sn}] (lejárat: [{expiryDate}])."
        " A Pro lejárat: [{expiryTm}], összesen [{days}] nap. Fizetendő: [{day}] nap díja.",

    # ---------- report ----------
    "report.bind.user.placeholder":
        "Adja meg a felhasználó azonosítóját (ID)",
    "report.bind.user.placeholder.age.tips": "Adja meg a felhasználó életkorát",
    "report.bind.user.placeholder.error":
        "Feladat hozzáadva. Lásd a jelentéslistán.",
    "report.bind.user.placeholder.error.duplicate":
        "A feladat már kötött, ne kösse újra.",
    "report.bind.user.placeholder.height.tips": "Adja meg a magasságát",
    "report.bind.user.placeholder.height.tips.placeholder":
        "Adja meg testmagasságát (70–205 cm)",
    "report.bind.user.placeholder.id.tips": "Adja meg a felhasználó azonosítóját",
    "report.bind.user.placeholder.id.tips.error":
        "Helyes azonosító (pl. betegkártya, szem. ig.)",
    "report.bind.user.placeholder.id.tips.error.placeholder":
        "Felhasználói azonosító (pontosan adja meg!)",
    "report.bind.user.placeholder.id.tips.placeholder":
        "Azonosító: betegkártya / személyi igazolvány",
    "report.bind.user.placeholder.name.tips": "Adja meg a felhasználó nevét",
    "report.bind.user.placeholder.new":
        "Új felhasználói azonosító – ellenőrizze!",
    "report.bind.user.placeholder.old":
        "Töltse ki pontosan az adatokat!",
    "report.bind.user.placeholder.success":
        "Kötés előtt ellenőrizze az adatokat!",
    "report.bind.user.placeholder.tip":
        "WeChat és vendégfiók: különálló rendszerek.",
    "report.common.delete.tip":
        "Biztosan törli a mérési jelentést? A törölt adatok nem állíthatók vissza!",
    "report.hip.contrast.tip":
        "*A ruházat befolyásolja az eredményt. Pontos méréshez viseljen szűk ruhát.",
    "report.hip.contrast.tip.title": "*Csak tájékoztató jellegű",
    "report.list.editor.title": "Név szerkesztése",
    "report.nutrition.nutrition.folicAcid.title": "Folsav (per 100 g)",
    "report.nutrition.nutrition.title": "Élelmiszerek tápanyagai",
    "report.preview.tips2":
        "Nincs előnézet, próbálja újra a szintézis befejezése után!",
    "report.weight.info.title": "Súlypontellenőrzés",

    # ---------- role ----------
    "role.create.des.placeholder":
        "Jogosultságbeli különbségek vagy megjegyzések",
    "role.create.name.placeholder":
        "Pl. főedző, üzletvezető (munkakör megadása)",
    "role.create.name.placeholder1":
        "Jogosultság neve vagy üzleti jellemzője",
    "role.create.permission.title": "Új jogosultság",
    "role.edit.permission.title": "Jogosultság szerkesztése",
    "role.operate.disable.title": "Szerepkör használatban",

    # ---------- sign ----------
    "sign.input.tip3":
        "{name1} és {name2} pénzneme eltér; egy ügyfél csak egy valutát használhat.",
    "sign.input.tip4":
        "A(z) {name} mód már beállítva; típusonként csak egy fizetési mód engedélyezett.",
    "sign.result1.contact.title": "Értesítés nem érkezett?",
    "sign.result1.time":
        "Automatikus ellenőrzés: Ha az információ megfelel a szabványoknak, perceken belül"
        " elkészül (frissítse az oldalt vagy ellenőrizze az eredményt"
        " előregisztrált telefonján/e-mailjén)",
    "sign.tip":
        "*Töltse ki az adatokat és küldje be. Jóváhagyás után aktiváljuk a platformot.",

    # ---------- store ----------
    "store.account.placeholder": "Adja meg a WellnessHub felhasználónevét",
    "store.contact_phone.placeholder": "Adja meg a kapcsolattartó számát",
    "store.hotline.placeholder": "Adja meg a tanácsadási telefonszámát",
    "store.logo.tips":
        "A LOGO a feltöltés után megjelenik a tesztjelentésben; törléskor eltűnik.",
    "store.logo.tips1":
        'Töltsön fel 320×160 pixel méretű, 200kb alatti ".png" formátumú képet',
    "store.logo.tips2":
        "Bal felső sarokban jelenik meg; törléskor az alapértelmezett logó visszaáll.",
    "store.logo.tips3":
        "Megjegyzés: A beállítás után a bolt logója a konfigurált képként jelenik meg.",
    "store.logo.tips4":
        "Megjegyzés: A beállítás után a jelentés logója a konfigurált kép lesz.",
    "store.logo.tips5":
        "Feltöltés: logó a tesztjelentésen. Törlés: alap logó visszaáll.",
    "store.report.2023.desc":
        "Egészségügyi szabványokon alapuló mérési rendszer, naprakész értékekkel,"
        " ázsiai felhasználókra optimalizálva.",
    "store.report.showAiContent.title": "AI-tartalom megtekintése",
    "store.resetpwd.tip":
        "Megjegyzés: Az új jelszó csak a [Megerősít] gombra kattintáskor aktiválódik!",
    "store.upload.excel.nutritionalList.tips":
        "Excel feltöltése után az AI személyre szabott tápanyag-kiegészítőket javasol.",
    "store.upload.excel.tips":
        "Excel feltöltése után az AI edzéstervet generál a konfigurált eszközök alapján.",

    # ---------- aireport (buttons/errors not in initial 140) ----------
    "aireport.abnormal.btn.showaireport": "Jelentés",
    "aireport.abnormal.btn.viewoutline": "Terv mutatása",
    "aireport.abnormal.error4":
        "AI-elemzéshez szükséges: testösszetétel/testalak/testkerület.",
    "aireport.abnormal.error7":
        "AI-elemzés nem elérhető. Végezzen el egy testmérési csoportot is.",
    "aireport.abnormal.error8":
        "AI-elemzés nem elérhető. Végezzen el egy testmérési csoportot is.",

    # ---------- device (additional) ----------
    "device.renew.form.tip4.title": "Megjegyzés: Fizetendő ár",

    # ---------- forget (button) ----------
    "forget.pwd.button": "Jelszó küldése",

    # ---------- message (additional placeholders) ----------
    "message.menu.training.question.postpartum.day.placeholder22":
        "D) Erőteljes mozgás (pl. profi sportolók)",
    "message.menu.training.question.postpartum.day.placeholder42":
        "Fő mozgásforma (több is kiválasztható):",
    "message.menu.training.question.postpartum.day.placeholder45":
        "C) Intenzív mozgás: futás, úszás, aerobik",

    # ---------- report (additional nutrition titles) ----------
    "report.nutrition.nutrition.calcium.title": "Kalciumtartalom (100 g)",
    "report.nutrition.nutrition.iron.title": "Vastartalom (100 g)",

    # ---------- store (additional) ----------
    "store.report.biaStandard.tip3":
        "Csak Visbody-Rpro5, Dpro3, E260, Apro6 eszközökre érvényes;",
    "store.report.watermark.title": "Visbody vízjeles jelentés",

    # ---------- user ----------
    "user.health.record.additionalInfo.placeholder":
        "Edzéstapasztalat, betegségek, egyéb igények",
    "user.health.record.allergyHistoryDetail.placeholder":
        "Kérjük, adja meg az allergiás előzményeket.",
    "user.health.record.bodyShapeGoalsOther.placeholder":
        "Pl. egy meghatározott testterület vonalai",
    "user.health.record.currentMedication.placeholder":
        "Jelenlegi gyógyszerek megadása",
    "user.health.record.exerciseContraindications.placeholder":
        "Pl. kerülendő: dőlés, intenzív aerob edzés",
    "user.health.record.jointInjuryDetail.placeholder":
        "Sérült ízület helye, pl.: térd, váll",
    "user.health.record.trainingGoalSculptingDetail.placeholder":
        "Alakítandó testrész (pl. kar, has)",
}


# =========================================================
# Excel 样式辅助
# =========================================================
def _header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor="1F4E79"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }

def _ok_fill():
    return PatternFill("solid", fgColor="E2EFDA")   # 浅绿

def _warn_fill():
    return PatternFill("solid", fgColor="FFF2CC")   # 浅黄

def _over_fill():
    return PatternFill("solid", fgColor="FFDCDC")   # 浅红

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _wrap():
    return Alignment(vertical="top", wrap_text=True)


# =========================================================
# 主流程
# =========================================================
def main() -> None:
    csv_in = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Desktop" / "工作簿1.csv"
    xlsx_out = (
        Path(sys.argv[2]) if len(sys.argv) > 2
        else csv_in.parent / "匈牙利语翻译压缩稿.xlsx"
    )

    if not csv_in.exists():
        sys.exit(f"[错误] 文件不存在: {csv_in}")

    with open(csv_in, "r", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    # -------- 构建数据 --------
    records = []
    for r in rows[1:]:
        if len(r) < 3:
            continue
        key    = r[0].strip()
        old_hu = r[1].strip()
        new_hu = r[2].strip()
        status = r[3].strip() if len(r) > 3 else ""

        cat, max_len = classify(key)
        cur_len = len(new_hu)
        needs = (status == "变长") and bool(new_hu) and cur_len > max_len

        if key in COMPRESSED:
            final = COMPRESSED[key]
        else:
            final = new_hu   # 不需压缩，直接用 C 列新译文

        final_len = len(final)
        within = final_len <= max_len

        records.append(dict(
            key=key, old_hu=old_hu, new_hu=new_hu, status=status,
            cat=cat, max_len=max_len, cur_len=cur_len,
            needs="Y" if needs else "N",
            final=final, final_len=final_len,
            within="Y" if within else "N",
        ))

    # -------- 写 Excel --------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "翻译压缩稿"

    headers = [
        "key", "旧译文 (B列)", "新译文 (C列)", "变长状态",
        "场景类型", "字符上限", "当前字符数",
        "需压缩", "压缩后译文 [最终使用]", "压缩后字符数", "是否达标",
    ]
    col_widths = [38, 40, 40, 8, 10, 8, 8, 8, 40, 10, 8]

    hs = _header_style()
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hs["font"]
        cell.fill = hs["fill"]
        cell.alignment = hs["alignment"]
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28

    bd = _thin_border()
    wrap = _wrap()

    for ri, rec in enumerate(records, 2):
        vals = [
            rec["key"], rec["old_hu"], rec["new_hu"], rec["status"],
            rec["cat"], rec["max_len"], rec["cur_len"],
            rec["needs"], rec["final"], rec["final_len"], rec["within"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = bd
            cell.alignment = wrap

        # 行着色：仅对需压缩的行着色
        if rec["needs"] == "Y":
            fill = _ok_fill() if rec["within"] == "Y" else _over_fill()
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = fill

    # 冻结首行
    ws.freeze_panes = "A2"

    # 汇总 sheet
    ws2 = wb.create_sheet("汇总说明")
    total = len(records)
    compressed_count = sum(1 for r in records if r["needs"] == "Y")
    ok_count = sum(1 for r in records if r["needs"] == "Y" and r["within"] == "Y")
    over_count = compressed_count - ok_count
    ws2.append(["说明", "数值"])
    ws2.append(["总条目数", total])
    ws2.append(["需压缩条目数", compressed_count])
    ws2.append(["压缩后达标", ok_count])
    ws2.append(["压缩后仍超限（人工复核）", over_count])
    ws2.append([])
    ws2.append(["场景字符上限说明", ""])
    ws2.append(["按钮", 15])
    ws2.append(["标题/标签", 25])
    ws2.append(["占位提示", 45])
    ws2.append(["错误消息", 65])
    ws2.append(["提示", 80])
    ws2.append(["描述", 120])
    ws2.append(["说明文本", 200])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    wb.save(xlsx_out)
    print(f"Excel 已生成：{xlsx_out}")
    print(f"总条目：{total}  |  需压缩：{compressed_count}  |  达标：{ok_count}  |  仍超限：{over_count}")


if __name__ == "__main__":
    main()
