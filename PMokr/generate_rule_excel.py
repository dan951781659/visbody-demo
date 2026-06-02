# -*- coding: utf-8 -*-
"""生成 AI报告解读-综合结论规则库 完整版 Excel 文件"""
import csv
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def create_excel():
    if not HAS_OPENPYXL:
        print("请先安装: pip install openpyxl")
        return False

    wb = Workbook()
    ws_rules = wb.active
    ws_rules.title = "综合结论规则"

    # 表头
    headers = ["规则ID", "综合结论名称", "所属维度", "优先级", "必要指标组合(AND)", "可选强化指标", 
               "排除条件", "覆盖规则ID", "证据指标", "结论描述模板"]
    for col, h in enumerate(headers, 1):
        cell = ws_rules.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # 读取 CSV 数据
    csv_path = os.path.join(os.path.dirname(__file__), "AI报告解读-综合结论规则库-完整版.csv")
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row_num, row in enumerate(reader, 2):
            ws_rules.cell(row=row_num, column=1, value=row.get("规则ID", ""))
            ws_rules.cell(row=row_num, column=2, value=row.get("综合结论名称", ""))
            ws_rules.cell(row=row_num, column=3, value=row.get("所属维度", ""))
            ws_rules.cell(row=row_num, column=4, value=row.get("优先级", ""))
            ws_rules.cell(row=row_num, column=5, value=row.get("必要指标组合(AND)", ""))
            ws_rules.cell(row=row_num, column=6, value=row.get("可选强化指标", ""))
            ws_rules.cell(row=row_num, column=7, value=row.get("排除条件", ""))
            ws_rules.cell(row=row_num, column=8, value=row.get("覆盖规则ID", ""))
            ws_rules.cell(row=row_num, column=9, value=row.get("证据指标", ""))
            ws_rules.cell(row=row_num, column=10, value=row.get("结论描述模板", ""))

    # 调整列宽
    for col in range(1, 11):
        ws_rules.column_dimensions[get_column_letter(col)].width = 18 if col <= 4 else 35

    # 创建指标参考表
    ws_ind = wb.create_sheet("指标清单参考", 1)
    ind_headers = ["功能模块", "指标ID", "指标名称", "典型SKU"]
    for col, h in enumerate(ind_headers, 1):
        cell = ws_ind.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    indicators = [
        ("体成分", "BC_01", "体重", "全系列"),
        ("体成分", "BC_02", "去脂体重", "全系列"),
        ("体成分", "BC_03", "肌肉量", "M30/T50/T55等"),
        ("体成分", "BC_04", "体脂肪", "全系列"),
        ("体成分", "BC_05", "体脂率", "全系列"),
        ("体成分", "BC_06", "骨骼肌", "全系列"),
        ("体成分", "BC_07", "蛋白质", "全系列"),
        ("体成分", "BC_08", "无机盐", "M30等"),
        ("体成分", "BC_09", "总水分", "全系列"),
        ("体成分", "BC_10", "细胞内液", "VR Explorer/M30等"),
        ("体成分", "BC_11", "细胞外液", "同上"),
        ("体成分", "BC_12", "腰臀比", "全系列"),
        ("体成分", "BC_13", "基础代谢", "全系列"),
        ("体成分", "BC_14", "代谢年龄", "M30等"),
        ("体成分", "BC_15", "BMI", "全系列"),
        ("体成分", "BC_16", "内脏脂肪面积", "VR PRO5/S20等"),
        ("体成分", "BC_17", "内脏脂肪等级", "M30等"),
        ("体成分", "BC_18", "节段肌肉", "M30等"),
        ("体成分", "BC_19", "节段脂肪", "M30等"),
        ("体态评估", "PT_01", "头前引", "M30/S20/T50/T55等"),
        ("体态评估", "PT_02", "头侧歪", "同上"),
        ("体态评估", "PT_03", "高低肩", "同上"),
        ("体态评估", "PT_04", "左圆肩", "同上"),
        ("体态评估", "PT_05", "右圆肩", "同上"),
        ("体态评估", "PT_06", "骨盆前移", "同上"),
        ("体态评估", "PT_07", "骨盆前/后倾", "VR Explorer/VR PRO5等"),
        ("体态评估", "PT_08", "左膝评估", "M30/S20等"),
        ("体态评估", "PT_09", "右膝评估", "同上"),
        ("体态评估", "PT_10", "腿型(X/O/D/K)", "同上"),
        ("体态评估", "PT_11", "长短腿", "VR Explorer/VR PRO5等"),
        ("体态评估", "PT_12", "身体倾斜", "VR PRO5/VA PRO5等"),
        ("体态评估", "PT_13", "重心平衡", "同上"),
        ("体围", "BW_01", "颈围", "M30/VA PRO5/S20等"),
        ("体围", "BW_02~04", "上臂围/胸围", "同上"),
        ("体围", "BW_05~10", "大腿/小腿围", "同上"),
        ("体围", "BW_11~14", "腰围/臀围", "同上"),
        ("腰腹围度", "WA_01~12", "高腰/低腰/体积等", "VR Explorer/T50/T55等"),
        ("肩部功能", "SH_01~04", "外展/前屈上举", "VD PRO3/VE560/VA PRO5等"),
        ("髋关节", "HP_01", "左髋外展", "VR Explorer/M30等"),
        ("髋关节", "HP_02", "右髋外展", "同上"),
        ("髋关节", "HP_03", "左髋内收", "同上"),
        ("髋关节", "HP_04", "右髋内收", "同上"),
        ("脊柱评估", "SP_01", "脊柱关键曲度", "VR Explorer/M30等"),
        ("脊柱评估", "SP_02~16", "结构异常/疼痛风险等", "T50/T55等"),
        ("颈部功能", "NC_01~06", "颈椎6向活动度", "VR Explorer/M30等"),
        ("臀型", "BT_01", "臀型评估结果", "M30/VA PRO5等"),
        ("营养", "NU_01", "无电流营养分析", "M30/VR PRO5等"),
        ("FMS", "FMS_01~10", "深蹲/跨栏步等10项", "T50/T55"),
        ("体适能", "FT_01~10", "纵跳/核心/平衡等", "T50/T55"),
        ("青少年", "YG_01~03", "身高/BMI百分位等", "T50/T55"),
        ("平衡", "BL_01~28", "睁眼/闭眼/动态平衡等", "T50/T55"),
    ]
    for row_num, (mod, iid, name, sku) in enumerate(indicators, 2):
        ws_ind.cell(row=row_num, column=1, value=mod)
        ws_ind.cell(row=row_num, column=2, value=iid)
        ws_ind.cell(row=row_num, column=3, value=name)
        ws_ind.cell(row=row_num, column=4, value=sku)

    for col in range(1, 5):
        ws_ind.column_dimensions[get_column_letter(col)].width = 20

    out_path = os.path.join(os.path.dirname(__file__), "AI报告解读-综合结论规则库-完整版.xlsx")
    wb.save(out_path)
    print(f"已生成: {out_path}")
    return True

if __name__ == "__main__":
    create_excel()
